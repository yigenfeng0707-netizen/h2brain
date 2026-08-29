"""H2Brain - Accuracy Validation Module

Validates automatic trip segmentation and work-condition classification against
the official manual record spreadsheet ("测试车辆2#手工行程及工况记录表.xlsx")
provided in the T05 data package.

This module serves as ground-truth verification: the manual spreadsheet contains
human-verified trip boundaries, mileage, and work-condition categories for
Vehicle 2 (测试车辆2#, Xinjiang Turpan region). By comparing our algorithmic
output with these records, we quantify segmentation accuracy, mileage deviation,
and work-condition classification agreement.

References:
  - Manual record source: T05 data package, "测试车辆2#手工行程及工况记录表.xlsx"
  - Automatic segmentation: data_processor._segment_trips() (time-gap > 10 min)
  - Work-condition classification: data_processor._classify_road_condition()
"""

from __future__ import annotations

import logging
from datetime import datetime, time, timedelta
from typing import Any

import openpyxl

from .data_processor import DATA_DIR, VEHICLE_FILES, get_processor

logger = logging.getLogger("h2brain.validation")

# Path to the manual record spreadsheet
MANUAL_RECORD_FILE = DATA_DIR / "测试车辆2#手工行程及工况记录表.xlsx"


def _load_manual_records() -> list[dict[str, Any]]:
    """Load and parse the manual trip record spreadsheet for Vehicle 2.

    The spreadsheet has a 3-row header (merged cells, units row, data starts row 4).
    Key columns:
      - Col A: date
      - Col B/C: start/end time (formula references, may not resolve)
      - Col E: trip mileage (km)
      - Col F: total mileage (km)
      - Col I: weather
      - Col K: load status (车头/带挂)
      - Col L: total weight (tons)
      - Col O: start location
      - Col Z: end location
      - Col AI: work condition category (综合工况/国道工况/山区国道)

    Returns list of dicts with parsed trip records.
    """
    if not MANUAL_RECORD_FILE.exists():
        logger.warning("Manual record file not found: %s", MANUAL_RECORD_FILE)
        return []

    wb = openpyxl.load_workbook(MANUAL_RECORD_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]

    def _safe_float(val: Any, default: float | None = None) -> float | None:
        """Safely convert an openpyxl cell value to float."""
        if val is None:
            return default
        try:
            return float(val)
        except (TypeError, ValueError):
            return default

    records = []
    # Data starts at row 4 (row 1=title, row 2=column names, row 3=units)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        # Skip empty rows
        if not row[0]:
            continue

        date_val = row[0]
        trip_mileage = row[4]
        total_mileage = row[5]
        weather = row[8]
        load_status = row[10]
        total_weight = row[11]
        start_location = row[14]
        start_mileage = row[17]
        end_location = row[25]
        end_mileage = row[29]
        work_condition = row[34] if len(row) > 34 else None

        # Parse start/end times from columns N/Y (indices 13/24) which are time objects
        start_time = row[13] if isinstance(row[13], time) else None
        end_time = row[24] if isinstance(row[24], time) else None

        # Build datetime
        base_date = date_val if isinstance(date_val, datetime) else None
        start_dt = None
        end_dt = None
        if base_date and start_time:
            start_dt = datetime.combine(base_date.date(), start_time)
        if base_date and end_time:
            end_dt = datetime.combine(base_date.date(), end_time)
            # Handle overnight trips (end time < start time means next day)
            if start_dt and end_dt < start_dt:
                end_dt += timedelta(days=1)

        records.append(
            {
                "date": base_date.date().isoformat() if base_date else None,
                "start_datetime": start_dt.isoformat() if start_dt else None,
                "end_datetime": end_dt.isoformat() if end_dt else None,
                "trip_mileage_km": _safe_float(trip_mileage, 0.0),
                "total_mileage_km": _safe_float(total_mileage, 0.0),
                "weather": weather,
                "load_status": load_status,
                "total_weight_t": _safe_float(total_weight),
                "start_location": start_location,
                "start_mileage": _safe_float(start_mileage),
                "end_location": end_location,
                "end_mileage": _safe_float(end_mileage),
                "work_condition": work_condition,
            }
        )

    logger.info("Loaded %d manual trip records", len(records))
    return records


# Mapping from manual (official) work-condition categories to our road-condition types
# v2.1: 工程类别已对齐官方四分类 (怠速/场区-装卸 归入 综合工况), 直接映射
CONDITION_MAP = {
    "综合工况": ["综合工况", "怠速", "场区-装卸"],
    "国道工况": ["国道工况"],
    "山区国道": ["山区国道"],
    "平原高速": ["平原高速"],
}


def _match_work_condition(manual_category: str, auto_road_types: list[str]) -> bool:
    """Check if automatic road types are compatible with manual work condition."""
    if not manual_category or not auto_road_types:
        return False
    expected = CONDITION_MAP.get(manual_category, [])
    return any(rt in expected for rt in auto_road_types)


def validate_trip_segmentation(vehicle_id: str = "V2") -> dict[str, Any]:
    """Validate automatic trip segmentation against manual records.

    Compares:
    1. Trip count: manual vs automatic
    2. Total mileage: manual sum vs automatic sum
    3. Per-trip mileage: nearest-match deviation
    4. Work-condition classification agreement rate

    Returns a structured validation report.
    """
    manual_records = _load_manual_records()
    if not manual_records:
        return {"error": "Manual records not available", "vehicle_id": vehicle_id}

    processor = get_processor()
    auto_trips = processor.get_trips(vehicle_id)
    if not auto_trips:
        return {
            "error": f"No automatic trips for {vehicle_id}",
            "vehicle_id": vehicle_id,
        }

    # --- 1. Trip count comparison ---
    manual_count = len(manual_records)
    auto_count = len(auto_trips)

    # Filter out zero-mileage manual records (parking/idling entries)
    manual_moving = [r for r in manual_records if r["trip_mileage_km"] > 0]
    manual_moving_count = len(manual_moving)

    # --- 2. Total mileage comparison ---
    manual_total_mileage = sum(r["trip_mileage_km"] for r in manual_records)
    auto_total_mileage = sum(t["distance_km"] for t in auto_trips)
    mileage_deviation = abs(auto_total_mileage - manual_total_mileage)
    mileage_deviation_pct = (
        (mileage_deviation / manual_total_mileage * 100)
        if manual_total_mileage > 0
        else 0
    )

    # --- 2b. Daily mileage comparison (more fair when segmentation granularity differs) ---
    manual_by_date = {}
    for r in manual_records:
        d = r.get("date")
        if d:
            manual_by_date[d] = manual_by_date.get(d, 0) + r["trip_mileage_km"]

    auto_by_date = {}
    for t in auto_trips:
        d = t.get("date", "")
        if d:
            auto_by_date[d] = auto_by_date.get(d, 0) + t["distance_km"]

    daily_comparison = []
    all_dates = sorted(set(manual_by_date.keys()) | set(auto_by_date.keys()))
    for d in all_dates:
        m_km = manual_by_date.get(d, 0)
        a_km = auto_by_date.get(d, 0)
        daily_comparison.append(
            {
                "date": d,
                "manual_km": round(m_km, 1),
                "auto_km": round(a_km, 1),
                "diff_km": round(a_km - m_km, 1),
            }
        )

    # Daily MAPE (only for days with manual mileage > 10 km)
    significant_daily = [dc for dc in daily_comparison if dc["manual_km"] > 10]
    daily_mape = (
        sum(abs(dc["diff_km"]) / dc["manual_km"] * 100 for dc in significant_daily)
        / len(significant_daily)
        if significant_daily
        else 0
    )

    # --- 2c. Sensor cumulative mileage comparison ---
    # Manual records have total mileage (odometer) at start and end
    manual_start_km = (
        manual_records[0].get("total_mileage_km", 0) if manual_records else 0
    )
    manual_end_km = (
        manual_records[-1].get("total_mileage_km", 0) if manual_records else 0
    )
    manual_odometer_distance = manual_end_km - manual_start_km

    # --- 3. Per-trip mileage matching ---
    # Match each manual trip to the closest automatic trip by mileage
    auto_mileages = sorted([t["distance_km"] for t in auto_trips])
    manual_mileages = sorted([r["trip_mileage_km"] for r in manual_moving])

    per_trip_errors = []
    for m_mileage in manual_mileages:
        if auto_mileages:
            closest = min(auto_mileages, key=lambda a: abs(a - m_mileage))
            err = abs(closest - m_mileage)
            per_trip_errors.append(
                {
                    "manual_km": round(m_mileage, 1),
                    "matched_auto_km": round(closest, 1),
                    "abs_error_km": round(err, 1),
                    "rel_error_pct": round(err / m_mileage * 100, 1)
                    if m_mileage > 0
                    else 0,
                }
            )

    # Compute MAPE for trips with mileage > 5 km (filter trivial entries)
    significant_errors = [e for e in per_trip_errors if e["manual_km"] > 5]
    mape = (
        sum(e["rel_error_pct"] for e in significant_errors) / len(significant_errors)
        if significant_errors
        else 0
    )

    # --- 4. Work-condition classification agreement ---
    # Sample up to 10 trips for work-condition comparison (full scan is too slow)
    condition_matches = 0
    condition_total = 0
    sample_size = min(10, len(auto_trips))
    # Sample evenly across the trip list
    step = max(1, len(auto_trips) // sample_size)
    sampled_trips = auto_trips[::step][:sample_size]
    for trip in sampled_trips:
        detail = processor.get_trip_detail(vehicle_id, trip["trip_id"])
        if detail and detail.get("road_segments"):
            auto_road_types = list(
                set(seg["road_type"] for seg in detail["road_segments"])
            )
            # Find the manual record with closest mileage
            if manual_moving:
                closest_manual = min(
                    manual_moving,
                    key=lambda m: abs(m["trip_mileage_km"] - trip["distance_km"]),
                )
                if closest_manual.get("work_condition"):
                    condition_total += 1
                    if _match_work_condition(
                        closest_manual["work_condition"], auto_road_types
                    ):
                        condition_matches += 1

    condition_agreement = (
        condition_matches / condition_total * 100 if condition_total > 0 else 0
    )

    # --- 5. Summary ---
    trip_count_match = (
        "exact"
        if auto_count == manual_count
        else "close"
        if abs(auto_count - manual_count) <= 3
        else "deviation"
    )

    return {
        "vehicle_id": vehicle_id,
        "vehicle_label": VEHICLE_FILES.get(vehicle_id, {}).get("label", vehicle_id),
        "validation_date": datetime.now().isoformat(),
        "manual_record_source": "测试车辆2#手工行程及工况记录表.xlsx (T05 official data package)",
        "trip_count": {
            "manual_total": manual_count,
            "manual_moving": manual_moving_count,
            "automatic": auto_count,
            "match": trip_count_match,
            "note": f"Manual records include {manual_count - manual_moving_count} zero-mileage entries (parking/idling)",
        },
        "total_mileage": {
            "manual_trip_sum_km": round(manual_total_mileage, 1),
            "manual_odometer_km": round(manual_odometer_distance, 1),
            "automatic_km": round(auto_total_mileage, 1),
            "deviation_vs_trip_sum_pct": round(mileage_deviation_pct, 1),
            "deviation_vs_odometer_pct": round(
                abs(auto_total_mileage - manual_odometer_distance)
                / manual_odometer_distance
                * 100
                if manual_odometer_distance > 0
                else 0,
                1,
            ),
            "note": "Manual trip sum may differ from odometer due to excluded zero-mileage entries",
        },
        "daily_mileage": {
            "days_compared": len(significant_daily),
            "daily_mape_pct": round(daily_mape, 1),
            "details": daily_comparison,
        },
        "per_trip_accuracy": {
            "matched_trips": len(per_trip_errors),
            "significant_trips": len(significant_errors),
            "mape_pct": round(mape, 1),
            "details": per_trip_errors[:20],  # Show first 20 for brevity
        },
        "work_condition_agreement": {
            "matched": condition_matches,
            "total_compared": condition_total,
            "agreement_rate_pct": round(condition_agreement, 1),
            "note": "Compares auto road-type classification with manual 综合工况/国道工况/山区国道 categories",
        },
        "overall_assessment": _assess_overall(
            trip_count_match,
            mileage_deviation_pct,
            mape,
            condition_agreement,
            daily_mape,
        ),
    }


def _assess_overall(
    trip_match: str,
    mileage_dev_pct: float,
    mape: float,
    condition_agreement: float,
    daily_mape: float = 0,
) -> str:
    """Generate an overall assessment string."""
    scores = []
    scores.append(
        1.0 if trip_match == "exact" else 0.7 if trip_match == "close" else 0.4
    )
    scores.append(max(0, 1 - mileage_dev_pct / 30))
    scores.append(max(0, 1 - mape / 50))
    scores.append(condition_agreement / 100)
    scores.append(max(0, 1 - daily_mape / 30))
    avg = sum(scores) / len(scores)
    if avg >= 0.85:
        return f"Excellent (score={avg:.2f}): Algorithm output closely matches manual ground truth."
    elif avg >= 0.65:
        return f"Good (score={avg:.2f}): Algorithm captures the main trip structure with acceptable deviation."
    else:
        return f"Needs improvement (score={avg:.2f}): Significant deviation from manual records detected."
